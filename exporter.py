import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, CSV_COLUMNS, HOTELS_BY_ID

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(name="Calibri", bold=True, color="F0A500", size=11)

ROW_FILL_DEFAULT = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_ALT = PatternFill("solid", fgColor="F7F7F7")
ROW_FILL_NO_RESTAURANT = PatternFill("solid", fgColor="FFF8E1")
ROW_FILL_ERROR = PatternFill("solid", fgColor="FFEBEE")

ROW_FONT = Font(name="Calibri", size=10)
WRAP_FONT = Font(name="Calibri", size=10)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
)

COL_WIDTHS = {
    "Property ID": 14,
    "Does the hotel have an on site restaurant?": 20,
    "Restaurant Sequence": 12,
    "Restaurant Name": 28,
    "Restaurant Type": 20,
    "Phone number": 18,
    "Hours Of Operation": 36,
    "Cuisine Type": 18,
    "Dietary Menu Options": 22,
    "Meals Served": 22,
    "Restaurant description": 50,
    "Book a table": 40,
    "View menu": 40,
    "Visit Website": 40,
    "Dining Page Headline": 40,
    "Dining Page Description": 60,
    "Experience Page Dining Headline": 40,
    "Experience Page Dining Description": 60,
}

WRAP_COLUMNS = {
    "Hours Of Operation",
    "Restaurant description",
    "Dining Page Description",
    "Experience Page Dining Description",
}


def _write_headers(ws):
    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22


def _write_row(ws, row_idx: int, row_data: dict, fill: PatternFill):
    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        value = row_data.get(col_name, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.font = WRAP_FONT if col_name in WRAP_COLUMNS else ROW_FONT
        cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=(col_name in WRAP_COLUMNS),
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 60 if any(
        row_data.get(c, "") for c in WRAP_COLUMNS
    ) else 18


def _set_column_widths(ws):
    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)


def _pick_fill(row_data: dict, row_idx: int) -> PatternFill:
    if row_data.get("Does the hotel have an on site restaurant?", "").lower() == "no":
        return ROW_FILL_NO_RESTAURANT
    return ROW_FILL_DEFAULT if row_idx % 2 == 0 else ROW_FILL_ALT


def export_single(prop_id: str, rows: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    hotel = HOTELS_BY_ID.get(prop_id, {})
    hotel_name = hotel.get("Nome account", prop_id)

    wb = Workbook()
    ws = wb.active
    ws.title = prop_id
    ws.freeze_panes = "A2"

    _write_headers(ws)

    for i, row_data in enumerate(rows, start=2):
        fill = _pick_fill(row_data, i)
        _write_row(ws, i, row_data, fill)

    _set_column_widths(ws)

    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in hotel_name)[:40]
    filename = f"{prop_id}-{safe_name}-dining.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return path


def export_all(all_results: dict) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    done_results = {
        pid: data for pid, data in all_results.items()
        if data.get("status") == "done" and data.get("rows")
    }

    if not done_results:
        raise ValueError("No completed hotels to export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dining Data"
    ws.freeze_panes = "A2"

    _write_headers(ws)

    row_idx = 2
    for prop_id, data in done_results.items():
        for row_data in data["rows"]:
            fill = _pick_fill(row_data, row_idx)
            _write_row(ws, row_idx, row_data, fill)
            row_idx += 1

    _set_column_widths(ws)

    filename = f"bwh-dining-export-{date.today().isoformat()}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return path
