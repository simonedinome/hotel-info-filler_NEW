from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ALL_CATEGORIES, OUTPUT_DIR
from core.checkpoint import load_checkpoint
from schemas import load_schema


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1A1A2E")
HEADER_FONT = Font(color="F0A500", bold=True)
DONE_FILLS = [
    PatternFill(fill_type="solid", fgColor="FFFFFF"),
    PatternFill(fill_type="solid", fgColor="F5F5F5"),
]
DONE_EMPTY_FILL = PatternFill(fill_type="solid", fgColor="FFF3E0")
NO_DATA_FILL = PatternFill(fill_type="solid", fgColor="FFF8E1")
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFEBEE")


def _rows_for_export(checkpoint: dict) -> list[tuple[dict, dict]]:
    rows = []
    for entry in checkpoint.values():
        if entry.get("status") in {"done", "done_empty"}:
            for row in entry.get("rows", []):
                rows.append((entry, row))
    return rows


def _export_paths(category: str) -> tuple[Path, Path]:
    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_suffix = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = output_dir / f"bwh-{category}-export-{date_suffix}.xlsx"
    csv_path = output_dir / f"bwh-{category}-export-{date_suffix}.csv"
    return xlsx_path, csv_path


def export_category_csv(category: str) -> str:
    checkpoint = load_checkpoint(category)
    schema_module = load_schema(category)
    columns = [column["key"] for column in schema_module.COLUMNS]
    export_rows = _rows_for_export(checkpoint)
    _, csv_path = _export_paths(category)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for _, row_data in export_rows:
            writer.writerow({column: row_data.get(column) for column in columns})

    return str(csv_path)


def export_category(category: str) -> str:
    checkpoint = load_checkpoint(category)
    schema_module = load_schema(category)
    columns = [column["key"] for column in schema_module.COLUMNS]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = category
    sheet.freeze_panes = "A2"

    for column_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    export_rows = _rows_for_export(checkpoint)
    for row_index, (entry, row_data) in enumerate(export_rows, start=2):
        status = entry.get("status")
        if status == "done_empty":
            fill = DONE_EMPTY_FILL
        elif entry.get("has_category") is False:
            fill = NO_DATA_FILL
        else:
            fill = DONE_FILLS[(row_index - 2) % 2]

        for column_index, column_name in enumerate(columns, start=1):
            value = row_data.get(column_name)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column_index == 1 and entry.get("writer_failed"):
                cell.comment = Comment("Writer failed - editorial fields empty", "Codex")

    error_row_index = len(export_rows) + 2
    for prop_id, entry in checkpoint.items():
        if entry.get("status") != "error":
            continue
        sheet.cell(row=error_row_index, column=1, value=prop_id).fill = ERROR_FILL
        sheet.cell(row=error_row_index, column=2, value="ERROR").fill = ERROR_FILL
        error_row_index += 1

    for column_index in range(1, len(columns) + 1):
        letter = get_column_letter(column_index)
        width = 14
        for row in sheet.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                if cell.value is None:
                    continue
                width = min(60, max(width, len(str(cell.value)) + 2))
        sheet.column_dimensions[letter].width = width

    xlsx_path, _ = _export_paths(category)
    workbook.save(xlsx_path)
    export_category_csv(category)
    return str(xlsx_path)


def export_all() -> list[str]:
    return [export_category(category) for category in ALL_CATEGORIES]
