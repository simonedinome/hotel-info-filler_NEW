"""Standalone script: find golf clubs within 16 km of each hotel using
Google Places Nearby Search, then export results to Excel."""

from __future__ import annotations

import argparse
import math
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import GOOGLE_PLACES_API_KEY, OUTPUT_DIR, load_hotels

GOLF_RADIUS_M = 16_000  # 16 km
_NEARBY_URL = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
_DETAILS_URL = "https://maps.googleapis.com/maps/api/place/details/json"

# Place types that indicate it's a hotel/resort, not a dedicated golf club
_EXCLUDE_TYPES = {"lodging", "hotel", "resort"}

# Keywords in the name that disqualify a result (mini-golf, hotels, etc.)
_EXCLUDE_NAME_KEYWORDS = {
    "mini", "minigolf", "mini-golf", "pitch and putt", "pitch & putt",
    "hotel", "resort", "albergo", "relais", "spa", "agriturismo",
}

def _is_golf_club(place_name: str, place_types: list[str]) -> bool:
    """Return True only if the place is a dedicated golf club/course."""
    # Exclude if Google tagged it as lodging
    if any(t in _EXCLUDE_TYPES for t in place_types):
        return False
    # Exclude if the name contains disqualifying keywords
    name_lower = place_name.lower()
    if any(kw in name_lower for kw in _EXCLUDE_NAME_KEYWORDS):
        return False
    return True
_HEADER_BG = "1F3864"
_HEADER_FG = "FFD700"


# ---------------------------------------------------------------------------
# Google Places helpers
# ---------------------------------------------------------------------------

def _find_golf_clubs_nearby(lat: float, lon: float, api_key: str) -> list[dict]:
    """Call Nearby Search for golf_course, follow next_page_token (up to 3 pages)."""
    results: list[dict] = []
    params: dict = {
        "location": f"{lat},{lon}",
        "radius": GOLF_RADIUS_M,
        "type": "golf_course",
        "key": api_key,
    }
    for _ in range(3):
        resp = requests.get(_NEARBY_URL, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        results.extend(data.get("results", []))
        token = data.get("next_page_token")
        if not token:
            break
        # Google requires a short delay before the token becomes valid
        time.sleep(2)
        params = {"pagetoken": token, "key": api_key}
    return results


def _get_place_details(place_id: str, api_key: str) -> dict:
    """Fetch name, website, formatted_address, geometry, types from Place Details."""
    params = {
        "place_id": place_id,
        "fields": "name,website,formatted_address,geometry,types",
        "key": api_key,
    }
    resp = requests.get(_DETAILS_URL, params=params, timeout=15)
    resp.raise_for_status()
    return resp.json().get("result", {})


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6_371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))


_LAT_RANGE = (35.0, 48.0)   # Italy bounding box
_LON_RANGE = (5.0, 22.0)


def _parse_coord(value: str, lo: float, hi: float) -> float | None:
    """Parse a coordinate string within the expected [lo, hi] range.

    Handles three input formats produced by Salesforce → Excel exports:
      - Standard float string:   "43.78141" or "43,78141"
      - Raw integer string:      "43781410"  (no dots)
      - Italian thousands format: "4.378.141" or "43.781.410"
    """
    s = str(value).strip()
    if not s:
        return None

    # Remove Italian thousands separators (two or more dots = display formatting)
    if s.count(".") >= 2:
        s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        raw = float(s)
    except ValueError:
        return None

    if lo <= raw <= hi:
        return raw

    # Value is out of range — likely a raw integer needing division.
    # Try decreasing powers of 10 until the result falls in [lo, hi].
    if raw > hi and raw == int(raw):
        for exp in range(5, 8):
            result = raw / (10 ** exp)
            if lo <= result <= hi:
                return result

    return None


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def find_nearby_golf(hotel: dict, api_key: str) -> list[dict]:
    """Return a list of golf club dicts (sorted by distance) near *hotel*."""
    lat = _parse_coord(hotel.get("fLatitude", ""), *_LAT_RANGE)
    lon = _parse_coord(hotel.get("fLongitude", ""), *_LON_RANGE)
    if lat is None or lon is None:
        return []

    raw_results = _find_golf_clubs_nearby(lat, lon, api_key)

    clubs: list[dict] = []
    for place in raw_results:
        place_id = place.get("place_id", "")
        details = _get_place_details(place_id, api_key) if place_id else {}

        loc = details.get("geometry", {}).get("location", {})
        club_lat = loc.get("lat")
        club_lon = loc.get("lng")

        if club_lat is not None and club_lon is not None:
            dist = _haversine_km(lat, lon, club_lat, club_lon)
            if dist > GOLF_RADIUS_M / 1000:
                continue
        else:
            dist = None

        name = details.get("name") or place.get("name", "")
        types = details.get("types") or place.get("types") or []

        if not _is_golf_club(name, types):
            continue

        clubs.append({
            "name": name,
            "address": details.get("formatted_address", ""),
            "website": details.get("website", ""),
            "distance_km": round(dist, 2) if dist is not None else "",
        })

    clubs.sort(key=lambda c: c["distance_km"] if isinstance(c["distance_km"], float) else 999.0)
    return clubs


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

_COLUMNS = [
    "Property ID",
    "Hotel Name",
    "Latitudine",
    "Longitudine",
    "Golf Club Name",
    "Distance (km)",
    "Address",
    "Website",
]


def export_results(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Golf Proximity"

    header_font = Font(bold=True, color=_HEADER_FG)
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)

    for col_idx, col_name in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append([
            row["prop_id"],
            row["hotel_name"],
            row["lat"],
            row["lon"],
            row["club_name"],
            row["distance_km"],
            row["address"],
            row["website"],
        ])

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Find golf clubs within 16 km of each hotel using Google Places API."
    )
    parser.add_argument("--property-id", help="Process only this Property ID")
    parser.add_argument("--output", help="Output Excel file path (default: output/golf-proximity-YYYY-MM-DD.xlsx)")
    args = parser.parse_args()

    if not GOOGLE_PLACES_API_KEY:
        sys.exit("Error: GOOGLE_PLACES_API_KEY is not set. Add it to .env and retry.")

    hotels = load_hotels()
    if args.property_id:
        hotels = [h for h in hotels if h["Property ID"] == args.property_id]
        if not hotels:
            sys.exit(f"Error: Property ID '{args.property_id}' not found in input/export-hotel.csv.")

    hotels_with_coords = [
        h for h in hotels if h.get("fLatitude") and h.get("fLongitude")
    ]

    if not hotels_with_coords:
        print(
            "No hotels with coordinates found.\n"
            "Add 'fLatitude' and 'fLongitude' columns to input/export-hotel.xlsx and retry."
        )
        return

    all_rows: list[dict] = []
    total = len(hotels_with_coords)

    for idx, hotel in enumerate(hotels_with_coords, 1):
        prop_id = hotel["Property ID"]
        name = hotel.get("Nome account") or prop_id
        lat_raw = hotel.get("fLatitude", "")
        lon_raw = hotel.get("fLongitude", "")
        print(f"[{idx}/{total}] {name}  ({lat_raw}, {lon_raw})")

        clubs = find_nearby_golf(hotel, GOOGLE_PLACES_API_KEY)
        print(f"  → {len(clubs)} golf club(s) found")

        if clubs:
            for club in clubs:
                all_rows.append({
                    "prop_id": prop_id,
                    "hotel_name": name,
                    "lat": lat_raw,
                    "lon": lon_raw,
                    "club_name": club["name"],
                    "distance_km": club["distance_km"],
                    "address": club["address"],
                    "website": club["website"],
                })
        else:
            all_rows.append({
                "prop_id": prop_id,
                "hotel_name": name,
                "lat": lat_raw,
                "lon": lon_raw,
                "club_name": "",
                "distance_km": "",
                "address": "",
                "website": "",
            })

        if idx < total:
            time.sleep(1)

    output_path = args.output or str(
        Path(OUTPUT_DIR) / f"golf-proximity-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    export_results(all_rows, output_path)
    print(f"\nSaved: {output_path}  ({len(all_rows)} rows)")


if __name__ == "__main__":
    main()
