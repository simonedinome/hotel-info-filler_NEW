from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook


def load_dotenv_file(path: str = ".env") -> None:
    env_path = Path(path)
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


load_dotenv_file()

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "")
FIRECRAWL_API_KEY = os.getenv("FIRECRAWL_API_KEY", "")
GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY", "")


#GOOGLE_API_KEY = "YOUR_KEY_HERE"
#OPENROUTER_API_KEY = "YOUR_KEY_HERE"

GEMINI_MODEL = "gemini-flash-latest"
OPENROUTER_MODEL = "openai/gpt-4o"

GEMINI_TEMPERATURE = 0.1
WRITER_TEMPERATURE = 0.4
GEMINI_MAX_TOKENS = 8192
WRITER_MAX_TOKENS = 4096

FIRECRAWL_MAX_DEPTH = 3
FIRECRAWL_PAGE_LIMIT = 50

MIN_CATEGORY_CHARS = 300
REQUEST_DELAY = 3
VERIFIER_RETRY_MAX = 3
WRITER_RETRY_MAX = 3

SEARCH_DOMAINS = [
    "{hotel_website}",
    "bestwestern.it",
    "book.bestwestern.it"
]

PAGES_DIR = "pages"
OUTPUT_DIR = "output"
PROMPTS_DIR = "prompts"
SCHEMAS_DIR = "schemas"
INPUT_DIR = "input"
HOTELS_INPUT_PATH = str(Path(INPUT_DIR) / "export-hotel.xlsx")

CATEGORY_KEYWORDS = {
    "dining": [
        "restaurant",
        "bar",
        "dining",
        "food",
        "breakfast",
        "cuisine",
        "menu",
        "bistro",
        "cafe",
        "buffet",
        "lounge",
        "brasserie",
        "brunch",
        "lunch",
        "dinner",
        "ristorante",
        "colazione",
        "cucina",
        "pranzo",
        "cena",
        "tavolo",
    ],
    "spa": [
        "spa",
        "wellness",
        "massage",
        "treatment",
        "sauna",
        "hammam",
        "thermal",
        "facial",
        "body wrap",
        "relaxation",
        "hydrotherapy",
        "steam room",
        "massaggio",
        "trattamento",
        "benessere",
        "terme",
        "relax",
    ],
    "pool": [
        "pool",
        "swim",
        "swimming",
        "aqua",
        "water",
        "lazy river",
        "waterpark",
        "rooftop pool",
        "resort pool",
        "splash",
        "piscina",
        "nuoto",
        "vasca",
        "acquapark",
    ],
    "meetings": [
        "meeting",
        "event",
        "conference",
        "wedding",
        "banquet",
        "congress",
        "ballroom",
        "boardroom",
        "seminar",
        "co-working",
        "coworking",
        "venue",
        "sala",
        "riunione",
        "matrimonio",
        "evento",
        "convegno",
        "banchetto",
    ],
    "golf": [
        "golf",
        "course",
        "green",
        "driving range",
        "clubhouse",
        "fairway",
        "pro shop",
        "caddy",
        "tee",
        "bunker",
        "buca",
        "campo da golf",
        "campo golf",
    ],
    "experiences": [
        "beach",
        "casino",
        "marina",
        "fishing",
        "entertainment",
        "rooftop",
        "stargazing",
        "live music",
        "tennis",
        "basketball",
        "volleyball",
        "culinary",
        "cooking class",
        "wine tasting",
        "spiaggia",
        "pesca",
        "intrattenimento",
        "terrazza",
    ],
}

ALL_CATEGORIES = ["dining", "spa", "pool", "meetings", "golf", "experiences"]


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def load_hotels(path: str | None = None) -> list[dict]:
    workbook_path = Path(path or HOTELS_INPUT_PATH)
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        workbook.close()
        return []
    indexes = {name: index for index, name in enumerate(header)}

    def cell(row_values, column_name: str):
        index = indexes.get(column_name)
        if index is None or index >= len(row_values):
            return ""
        return row_values[index]

    hotels = []
    for row in rows:
        prop_id = _normalize_cell(cell(row, "Property ID"))
        if not prop_id:
            continue
        hotels.append(
            {
                "Property ID": prop_id,
                "Nome account": _normalize_cell(cell(row, "Nome account")),
                "Sito Web": _normalize_cell(cell(row, "Sito Web")),
                "Latitudine": _normalize_cell(cell(row, "Latitudine")),
                "Longitudine": _normalize_cell(cell(row, "Longitudine")),
            }
        )
    workbook.close()
    return hotels


def hotels_by_id(path: str | None = None) -> dict[str, dict]:
    return {hotel["Property ID"]: hotel for hotel in load_hotels(path)}

def validate_api_keys() -> None:
    missing = []
    if not GOOGLE_API_KEY:
        missing.append("GOOGLE_API_KEY")
    if not OPENROUTER_API_KEY:
        missing.append("OPENROUTER_API_KEY")
    if missing:
        raise RuntimeError(f"Missing required environment variables: {', '.join(missing)}")
